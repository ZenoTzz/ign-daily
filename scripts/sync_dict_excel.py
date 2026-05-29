"""同步JSON词库到Excel"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

JSON_PATH = r'C:\Users\Administrator\.openclaw\workspace\game_names_dict.json'
XLSX_PATH = r'X:\IGN_Daily_News\游戏影视名称词库.xlsx'

with open(JSON_PATH, 'r', encoding='utf-8') as f:
    d = json.load(f)

wb = Workbook()
wb.remove(wb.active)

# 表头样式
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

cat_names = {
    'games': '游戏',
    'movies_tv': '影视',
    'companies': '公司',
    'people': '人物',
    'media': '媒体',
    'terms': '术语',
}

for cat_key, cat_name in cat_names.items():
    if cat_key not in d or not d[cat_key]:
        continue
    ws = wb.create_sheet(cat_name)
    headers = ['English Name', '中文译名', 'Source']
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center
    
    items = d[cat_key]
    row = 2
    for en, info in sorted(items.items()):
        if isinstance(info, dict):
            cn = info.get('cn', '')
            src = info.get('source', '')
        else:
            cn = str(info)
            src = ''
        ws.cell(row, 1, en).alignment = left_top
        ws.cell(row, 2, cn).alignment = left_top
        ws.cell(row, 3, src).alignment = left_top
        for col in range(1, 4):
            ws.cell(row, col).border = border
            ws.cell(row, col).font = Font(name='微软雅黑', size=10)
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.freeze_panes = 'A2'

wb.save(XLSX_PATH)
print(f'Saved: {XLSX_PATH}')
print(f'Size: {os.path.getsize(XLSX_PATH)} bytes')
